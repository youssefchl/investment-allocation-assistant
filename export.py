from __future__ import annotations

import os
from dataclasses import dataclass
from pathlib import Path

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from fpdf import FPDF
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

from .data import company_info


@dataclass(frozen=True)
class ExportPaths:
    excel_path: Path
    pdf_path: Path
    charts_dir: Path


def _format_excel(excel_path: Path) -> None:
    wb = load_workbook(excel_path)
    thin = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin
                if isinstance(cell.value, float):
                    cell.number_format = "0.00"

        for cell in ws[1]:
            cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")

        for col in ws.columns:
            max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col)
            ws.column_dimensions[col[0].column_letter].width = max(10, max_len + 2)

    wb.save(excel_path)


def export_excel(out_path: Path, base100: pd.DataFrame, beta: pd.Series, corr: pd.DataFrame, fiche: pd.DataFrame) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(out_path, engine="openpyxl") as w:
        base100.to_excel(w, sheet_name="Base100")
        beta.to_frame("Beta").to_excel(w, sheet_name="Beta")
        corr.to_excel(w, sheet_name="Correlation")
        fiche.to_excel(w, sheet_name="Fiche", index=False)
    _format_excel(out_path)


class PdfReport(FPDF):
    def footer(self) -> None:
        self.set_y(-15)
        self.set_font("Helvetica", "I", 8)
        self.set_text_color(120, 120, 120)
        self.cell(0, 10, f"Page {self.page_no()}", 0, 0, "C")


def _save_chart(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    plt.tight_layout()
    plt.savefig(path, dpi=200)
    plt.close()


def export_pdf(out_path: Path, sector: str, prices_daily: pd.DataFrame, base100: pd.DataFrame, corr: pd.DataFrame, allocation: pd.Series, amount: float) -> Path:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    charts_dir = out_path.parent / "charts"
    charts_dir.mkdir(parents=True, exist_ok=True)

    # Price chart
    plt.figure(figsize=(10, 4))
    for c in prices_daily.columns:
        plt.plot(prices_daily.index, prices_daily[c], label=c)
    plt.title(f"Prices - {sector}")
    plt.xlabel("Date")
    plt.ylabel("Close")
    plt.legend(fontsize=7)
    price_png = charts_dir / "prices.png"
    _save_chart(price_png)

    # Base100 chart
    plt.figure(figsize=(10, 4))
    for c in base100.columns:
        plt.plot(base100.index, base100[c], label=c)
    plt.title(f"Performance (Base100) - {sector}")
    plt.xlabel("Date")
    plt.ylabel("Index")
    plt.legend(fontsize=7)
    base_png = charts_dir / "base100.png"
    _save_chart(base_png)

    # Correlation heatmap (matplotlib only)
    plt.figure(figsize=(6, 5))
    plt.imshow(corr.values, aspect="auto")
    plt.xticks(range(len(corr.columns)), corr.columns, rotation=45, ha="right")
    plt.yticks(range(len(corr.index)), corr.index)
    plt.title("Correlation matrix")
    plt.colorbar()
    corr_png = charts_dir / "corr.png"
    _save_chart(corr_png)

    pdf = PdfReport()
    pdf.set_auto_page_break(auto=True, margin=15)

    pdf.add_page()
    pdf.set_font("Helvetica", "B", 20)
    pdf.cell(0, 10, "Financial Analysis Report", ln=True, align="C")
    pdf.set_font("Helvetica", "", 12)
    pdf.ln(2)
    pdf.cell(0, 10, f"Sector: {sector}", ln=True, align="C")
    pdf.ln(6)

    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(0, 8, "Charts", ln=True)
    pdf.image(str(price_png), w=180)
    pdf.ln(3)
    pdf.image(str(base_png), w=180)
    pdf.ln(6)
    pdf.image(str(corr_png), w=150)

    pdf.add_page()
    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(0, 8, "Suggested allocation (educational)", ln=True)
    pdf.set_font("Helvetica", "", 11)
    for t, w in allocation.sort_values(ascending=False).items():
        pdf.cell(0, 7, f"{t}: {w*100:.1f}%  (~{w*amount:.2f})", ln=True)

    pdf.output(str(out_path))
    return out_path


def build_fiche(tickers: list[str]) -> pd.DataFrame:
    rows = []
    for t in tickers:
        info = company_info(t)
        rows.append(
            {
                "Company": info.get("longName", t),
                "Ticker": t,
                "Industry": info.get("industry", "N/A"),
                "MarketCap ($B)": round((info.get("marketCap") or 0) / 1e9, 2) if info.get("marketCap") else np.nan,
                "Revenue ($B)": round((info.get("totalRevenue") or 0) / 1e9, 2) if info.get("totalRevenue") else np.nan,
                "PE": info.get("trailingPE", np.nan),
                "DividendYield (%)": round((info.get("dividendYield") or 0) * 100, 2) if info.get("dividendYield") else np.nan,
            }
        )
    return pd.DataFrame(rows)
